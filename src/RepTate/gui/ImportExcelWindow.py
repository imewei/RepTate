# RepTate: Rheology of Entangled Polymers: Toolkit for the Analysis of Theory and Experiments
# --------------------------------------------------------------------------------------------------------
#
# Authors:
#     Jorge Ramirez, jorge.ramirez@upm.es
#     Victor Boudara, victor.boudara@gmail.com
#
# Useful links:
#     http://blogs.upm.es/compsoftmatter/software/reptate/
#     https://github.com/jorge-ramirez-upm/RepTate
#     http://reptate.readthedocs.io
#
# --------------------------------------------------------------------------------------------------------
#
# Copyright (2017-2023): Jorge Ramirez, Victor Boudara, Universidad Politécnica de Madrid, University of Leeds
#
# This file is part of RepTate.
#
# RepTate is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# RepTate is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with RepTate.  If not, see <http://www.gnu.org/licenses/>.
#
# --------------------------------------------------------------------------------------------------------
"""Module for importing data form Excel spreadsheets

"""
import sys
import os
import numpy as np
from PySide6.QtUiTools import loadUiType
from PySide6.QtCore import Qt, QItemSelectionModel
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QTableWidgetItem,
    QTableWidget,
    QAbstractItemView,
    QMessageBox,
)
from openpyxl import load_workbook
from xlrd import open_workbook
import RepTate

if getattr(sys, "frozen", False):
    # If the application is run as a bundle, the PyInstaller bootloader
    # extends the sys module by a flag frozen=True and sets the app
    # path into variable _MEIPASS'.
    PATH = sys._MEIPASS
else:
    PATH = os.path.dirname(os.path.abspath(__file__))
Ui_ImportExcelMainWindow, QMainWindowImportExcel = loadUiType(
    os.path.join(PATH, "import_excel_dialog.ui")
)


class ImportExcelWindow(QMainWindowImportExcel, Ui_ImportExcelMainWindow):
    """Dialog window for importing data from Excel spreadsheets.

    This window provides a graphical interface for selecting Excel files (.xls or .xlsx),
    choosing which worksheet tab to import from, selecting data columns, and previewing
    the data before import. Supports both legacy .xls and modern .xlsx formats.

    The window handles:
        - File selection via file browser or drag-and-drop
        - Multiple worksheet tabs with preview
        - Column mapping to expected data types (x, y, z)
        - Row skipping for header lines
        - Data validation and NaN handling
        - File parameter extraction

    Attributes:
        list_AZ (list): Column labels from A to BZ for Excel column selection.
        MAX_ROW (int): Maximum number of rows to preview (100).
        MAX_COL (int): Maximum number of columns to preview.
        filepath (str): Full path to the selected Excel file.
        dir_start (str): Starting directory for file browser.
        is_xlsx (bool): True if file is .xlsx format, False if .xls.
        wb: Workbook object from openpyxl or xlrd.
        sheet: Current worksheet object.
        max_row (int): Maximum row in current sheet.
        max_col (int): Maximum column in current sheet.
        nskip (int): Number of header rows to skip.
        col_names (list): Expected column names from file type.
        col_units (list): Expected column units from file type.
        ncol (int): Number of expected data columns.
        file_param (list): List of file parameter names.
        qtables (dict): Maps sheet names to (QTableWidget, selected_columns) tuples.
        sheet_names (list): Names of all worksheets in the workbook.
    """
    list_AZ = [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
        "AA",
        "AB",
        "AC",
        "AD",
        "AE",
        "AF",
        "AG",
        "AH",
        "AI",
        "AJ",
        "AK",
        "AL",
        "AM",
        "AN",
        "AO",
        "AP",
        "AQ",
        "AR",
        "AS",
        "AT",
        "AU",
        "AV",
        "AW",
        "AX",
        "AY",
        "AZ",
        "BA",
        "BB",
        "BC",
        "BD",
        "BE",
        "BF",
        "BG",
        "BH",
        "BI",
        "BJ",
        "BK",
        "BL",
        "BM",
        "BN",
        "BO",
        "BP",
        "BQ",
        "BR",
        "BS",
        "BT",
        "BU",
        "BV",
        "BW",
        "BX",
        "BY",
        "BZ",
    ]
    MAX_ROW = 100
    MAX_COL = len(list_AZ)

    def __init__(self, parent=None, ftype=None):
        super().__init__()
        self.setupUi(self)
        # self.show()
        self.filepath = ""
        self.dir_start = os.path.join(RepTate.root_dir, "data")
        self.is_xlsx = True
        self.wb = None
        self.sheet = None
        self.max_row = 0
        self.nskip = 0
        self.max_col = 0
        self.select_file_tb.clicked.connect(self.handle_get_file)
        self.skip_sb.valueChanged.connect(self.handle_nskip_changed)
        self.qtabs.currentChanged.connect(self.handle_tab_changed)
        self.col1_cb.activated.connect(self.handle_col1_cb_activated)
        self.col2_cb.activated.connect(self.handle_col2_cb_activated)
        self.col3_cb.activated.connect(self.handle_col3_cb_activated)

        self.col_names = ftype.col_names
        self.col_units = ftype.col_units
        self.ncol = len(self.col_names)
        self.file_param = ftype.basic_file_parameters
        self.populate_file_param(self.file_param)
        self.update_cols_cb()

    def handle_col1_cb_activated(self):
        """Handle selection change in the first column combo box.

        Updates the selected column index for the first data column (typically x-values)
        and refreshes the preview table to highlight the newly selected column.
        """
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[0] = self.col1_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()

    def handle_col2_cb_activated(self):
        """Handle selection change in the second column combo box.

        Updates the selected column index for the second data column (typically y-values)
        and refreshes the preview table to highlight the newly selected column.
        """
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[1] = self.col2_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()

    def handle_col3_cb_activated(self):
        """Handle selection change in the third column combo box.

        Updates the selected column index for the third data column (typically z-values)
        and refreshes the preview table to highlight the newly selected column. Only
        active when the file type expects three or more data columns.
        """
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[2] = self.col3_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()

    def update_cols_cb(self):
        """Update column selection combo boxes based on file type requirements.

        Populates the column selection dropdowns with available Excel column labels
        (A, B, C, etc.) and sets their labels to indicate the expected data type
        (e.g., 'frequency [rad/s]'). For file types with only 2 columns, hides the
        third column selector.
        """
        self.col1_cb.clear()
        self.col2_cb.clear()
        self.col1.setText(
            "Select Column <b>%s [%s]</b>" % (self.col_names[0], self.col_units[0])
        )
        self.col2.setText(
            "Select Column <b>%s [%s]</b>" % (self.col_names[1], self.col_units[1])
        )
        self.col1_cb.addItems(self.list_AZ[: self.max_col])
        self.col2_cb.addItems(self.list_AZ[: self.max_col])
        self.col2_cb.setCurrentIndex(1)
        if self.ncol > 2:
            self.col3_cb.clear()
            self.col3.setText(
                "Select Column <b>%s [%s]</b>" % (self.col_names[2], self.col_units[2])
            )
            self.col3_cb.addItems(self.list_AZ[: self.max_col])
            self.col3_cb.setCurrentIndex(2)
        else:
            self.col3.hide()
            self.col3_cb.hide()

    def handle_tab_changed(self, idx):
        """Handle worksheet tab selection change.

        When the user switches between worksheet tabs, updates the column selection
        combo boxes to reflect the number of columns in the new sheet and restores
        the previously selected column indices for that sheet.

        Args:
            idx: Index of the newly selected tab.
        """
        table, selected_idx = self.qtables[self.qtabs.tabText(idx)]
        ncols = table.columnCount()
        self.col1_cb.clear()
        self.col2_cb.clear()
        self.col1_cb.addItems(self.list_AZ[:ncols])
        self.col2_cb.addItems(self.list_AZ[:ncols])
        self.col1_cb.setCurrentIndex(selected_idx[0])
        self.col2_cb.setCurrentIndex(selected_idx[1])
        if self.ncol > 2:
            self.col3_cb.clear()
            self.col3_cb.addItems(self.list_AZ[:ncols])
            self.col3_cb.setCurrentIndex(selected_idx[2])
        self.update_data_preview_table()

    def handle_nskip_changed(self):
        """Handle change in the number of rows to skip.

        Updates the preview table when the user changes the spin box value for
        skipping header rows. Skipped rows are not highlighted in the preview
        and will not be imported.
        """
        if self.wb == None:
            return
        self.nskip = self.skip_sb.value()
        self.update_data_preview_table()

    def col2num(self, col):
        """Convert Excel column letter(s) to numeric index.

        Converts column identifiers like 'A', 'B', 'AA', 'AB' to their numeric
        equivalents (1, 2, 27, 28) using base-26 arithmetic.

        Args:
            col: Excel column label (e.g., 'A', 'AB', 'BZ').

        Returns:
            Numeric column index (1-based).
        """
        num = 0
        for c in col:
            num = num * 26 + (ord(c) - ord("A")) + 1
        return num

    def update_data_preview_table(self):
        """Update the data preview table to highlight selected columns.

        Visually indicates which columns will be imported by:
            - Changing header labels from column letters to data type names
            - Selecting (highlighting) cells in the chosen columns
            - Excluding skipped rows from the selection

        The preview helps users verify they've selected the correct columns
        before importing.
        """
        idx = self.qtabs.currentIndex()
        sname = self.qtabs.tabText(idx)
        col1 = self.col2num(self.col1_cb.currentText()) - 1
        col2 = self.col2num(self.col2_cb.currentText()) - 1
        if self.ncol > 2:
            col3 = self.col2num(self.col3_cb.currentText()) - 1
        table, _ = self.qtables[sname]
        nrows = table.rowCount()
        ncols = table.columnCount()
        if (col1 < 0) or (col2 < 0) or (nrows == 0) or (ncols == 0):
            return
        header_labels = [self.list_AZ[i] for i in range(ncols)]
        header_labels[col1] = self.col_names[0]
        header_labels[col2] = self.col_names[1]
        indexes = [table.model().index(k, col1) for k in range(self.nskip, nrows)]
        indexes += [table.model().index(k, col2) for k in range(self.nskip, nrows)]
        if self.ncol > 2:
            indexes += [table.model().index(k, col3) for k in range(self.nskip, nrows)]
            header_labels[col3] = self.col_names[2]
        table.setHorizontalHeaderLabels(header_labels)
        flag = QItemSelectionModel.Select
        table.selectionModel().clearSelection()
        [table.selectionModel().select(i, flag) for i in indexes]
        table.setFocus()

    def get_data(self):
        """Extract and process data from the selected Excel worksheet.

        Reads data from the currently selected worksheet tab and columns,
        processes it (sorting by x-values, removing NaN x-values, optional
        interpolation for NaN y/z values), and returns it in a dictionary format.

        Returns:
            Dictionary containing:
                - error (bool): True if import failed, False otherwise.
                - errmsg (str): Error message if error is True.
                - file (str): Filename of the Excel file.
                - sheet (str): Name of the worksheet.
                - x (ndarray): First column data (sorted, NaN-free).
                - y (ndarray): Second column data (sorted, optionally interpolated).
                - z (ndarray): Third column data if ncol > 2 (sorted, optionally interpolated).
                - flag_nan (bool): True if any NaN values remain in y or z after processing.
                - col1, col2, col3 (str): Excel column letters used for each data column.
        """
        x = []
        y = []
        z = []
        if self.wb == None:
            msg = "Could not import data. Select an Excel file first."
            return {"error": True, "errmsg": msg}
        flag_nan = False
        col1 = self.col2num(self.col1_cb.currentText()) - 1
        col2 = self.col2num(self.col2_cb.currentText()) - 1
        if self.ncol > 2:
            col3 = self.col2num(self.col3_cb.currentText()) - 1
        sname = self.qtabs.tabText(self.qtabs.currentIndex())
        if self.is_xlsx:
            sheet = self.wb[sname]
            max_row = sheet.max_row
            max_col = sheet.max_column
        else:
            sheet = self.wb.sheet_by_name(sname)
            max_row = sheet.nrows
            max_col = sheet.ncols

        if max_col < min(3, self.ncol):
            # not enough data columns in the spreadsheet tab
            # min(3, ) as the Excel import is configured for 3 data columns max.
            msg = (
                "Could not import data. Need %d data columns and this spreadsheed has only %d column(s)"
                % (self.ncol, max_col)
            )
            return {"error": True, "errmsg": msg}

        for k in range(self.nskip, max_row):
            # x values
            if self.is_xlsx:
                cellx = sheet.cell(row=k + 1, column=col1 + 1)
            else:
                cellx = sheet.cell(k, col1)
            if hasattr(cellx, "value"):
                valx = cellx.value
            else:
                valx = ""
            try:
                x.append(float(valx))
            except (ValueError, TypeError):
                x.append(np.nan)
                flag_nan = True

            # y values
            if self.is_xlsx:
                celly = sheet.cell(row=k + 1, column=col2 + 1)
            else:
                celly = sheet.cell(k, col2)
            if hasattr(celly, "value"):
                valy = celly.value
            else:
                valy = ""
            try:
                y.append(float(valy))
            except (ValueError, TypeError):
                y.append(np.nan)
                flag_nan = True

            if len(self.col_names) > 2:
                # z values
                if self.is_xlsx:
                    cellz = sheet.cell(row=k + 1, column=col3 + 1)
                else:
                    cellz = sheet.cell(k, col3)
                if hasattr(cellz, "value"):
                    valz = cellz.value
                else:
                    valz = ""
                try:
                    z.append(float(valz))
                except (ValueError, TypeError):
                    z.append(np.nan)
                    flag_nan = True

        # Sort and clean data
        x = np.array(x)
        y = np.array(y)
        ind = np.argsort(x)
        x = x[ind]
        y = y[ind]
        ind2 = ~np.isnan(x)
        x = x[ind2]
        y = y[ind2]
        if self.cbInterpolate.isChecked():
            xynan = x[np.isnan(y)]
            xynotnan = x[~np.isnan(y)]
            ynotnan = y[~np.isnan(y)]
            yynan = np.interp(xynan, xynotnan, ynotnan, left=0, right=0)
            y[np.isnan(y)] = yynan
            flag_nan = False

        if len(self.col_names) > 2:
            z = np.array(z)
            z = z[ind]
            z = z[ind2]
            if self.cbInterpolate.isChecked():
                xznan = x[np.isnan(z)]
                xznotnan = x[~np.isnan(z)]
                znotnan = z[~np.isnan(z)]
                zznan = np.interp(xznan, xznotnan, znotnan, left=0, right=0)
                z[np.isnan(z)] = zznan

        res_dic = {
            "error": False,
            "file": self.selected_file_label.text(),
            "sheet": sname,
            "x": x,
            "y": y,
            "z": z,
            "flag_nan": flag_nan,
            "col1": self.col1_cb.currentText(),
            "col2": self.col2_cb.currentText(),
        }
        if len(self.col_names) > 2:
            res_dic["col3"] = self.col3_cb.currentText()
        return res_dic

    def populate_file_param(self, params):
        """Populate the file parameters text box with default values.

        Creates a semicolon-separated string of parameter=value pairs for all
        expected file parameters, initialized to 0. Users can edit these values
        before import.

        Args:
            params: List of parameter names expected by the file type.
        """
        self.file_param_txt.clear()
        txt = ""
        for p in params:
            txt += "%s=0;" % p
        self.file_param_txt.setText(txt)

    def handle_get_file(self):
        """Open file browser dialog to select an Excel file.

        Displays a file selection dialog filtered to show only .xls and .xlsx files.
        If a file is selected, delegates to handle_read_new_file to load and preview it.
        """
        # file browser window
        options = QFileDialog.Options()
        dilogue_name = "Select Excel Data File"
        ext_filter = "Excel file (*.xls *xlsx)"
        selected_file, _ = QFileDialog.getOpenFileName(
            self, dilogue_name, self.dir_start, ext_filter, options=options
        )
        self.handle_read_new_file(selected_file)

    def handle_read_new_file(self, path):
        """Load and display an Excel file with all its worksheets.

        Opens the Excel file (using openpyxl for .xlsx or xlrd for .xls), reads all
        worksheets, and creates preview tables for each. Creates a tab for each worksheet
        with a QTableWidget showing up to MAX_ROW rows and MAX_COL columns. Handles both
        modern (.xlsx) and legacy (.xls) Excel formats.

        Args:
            path: Full file path to the Excel file to load.
        """
        if not os.path.isfile(path):
            return
        self.dir_start = os.path.dirname(path)
        self.qtabs.blockSignals(True)
        self.clear_tabs()
        fname = os.path.basename(path)
        self.is_xlsx = os.path.splitext(path)[-1] == ".xlsx"
        self.selected_file_label.setText(fname)
        self.filepath = path
        try:
            if self.is_xlsx:
                self.wb = load_workbook(filename=self.filepath, data_only=True)
                self.sheet_names = self.wb.sheetnames
            else:
                self.wb = open_workbook(self.filepath)
                self.sheet_names = self.wb.sheet_names()
        except:
            # password protected?
            QMessageBox.warning(
                self, "Open Excel File", "Error: Could not read the Excel file."
            )
            return
        self.qtables = {}

        for sname in self.sheet_names:
            if self.is_xlsx:
                sheet = self.wb[sname]
                max_row = sheet.max_row
                max_col = sheet.max_column
            else:
                sheet = self.wb.sheet_by_name(sname)
                max_row = sheet.nrows
                max_col = sheet.ncols
            max_row = min(max_row, self.MAX_ROW)
            max_col = min(max_col, self.MAX_COL)
            qsheet = QTableWidget(max_row, max_col, self)
            qsheet.setSelectionMode(QAbstractItemView.NoSelection)
            for i in range(max_row):
                for j in range(max_col):
                    if self.is_xlsx:
                        cell = sheet.cell(row=i + 1, column=j + 1)
                    else:
                        cell = sheet.cell(i, j)
                    if hasattr(cell, "value"):
                        val = cell.value
                    else:
                        val = ""
                    item = QTableWidgetItem("%s" % val)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    qsheet.setItem(i, j, item)
            self.qtabs.addTab(qsheet, sname)
            selected_cols = [0, 1, 2]
            self.qtables[sname] = [qsheet, selected_cols]
        self.qtabs.blockSignals(False)
        self.qtabs.blockSignals(False)
        self.qtabs.setCurrentIndex(0)
        self.handle_tab_changed(0)

    def clear_tabs(self):
        """Remove all worksheet tabs and reset the interface.

        Clears all existing worksheet tabs from the tab widget, deletes their
        associated QTableWidget objects, resets the skip rows spin box to 0,
        and empties the qtables dictionary. Called before loading a new file.
        """
        for _ in range(self.qtabs.count()):
            w = self.qtabs.widget(0)
            self.qtabs.removeTab(0)
            del w
        self.qtables = {}
        self.skip_sb.blockSignals(True)
        self.skip_sb.setValue(0)
        self.skip_sb.blockSignals(False)
        self.nskip = 0

    def dragEnterEvent(self, e):
        """Handle drag enter events for file drag-and-drop.

        Accepts the drag event if it contains a file URI, enabling users to
        drag Excel files from their file manager directly onto the window.

        Args:
            e: QDragEnterEvent containing drag operation data.
        """
        if e.mimeData().hasFormat("text/uri-list"):
            e.accept()
        else:
            e.ignore()

    def dropEvent(self, e):
        """Handle drop events when files are dropped onto the window.

        Processes dropped files by checking if they have .xls or .xlsx extensions.
        If valid, loads the file using handle_read_new_file. Ignores files with
        other extensions.

        Args:
            e: QDropEvent containing the dropped file data.
        """
        path = e.mimeData().urls()[0].toLocalFile()
        if (
            os.path.splitext(path)[-1] == ".xls"
            or os.path.splitext(path)[-1] == ".xlsx"
        ):
            self.handle_read_new_file(path)
        else:
            pass
            # print("not a readable file")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    GUI = Window()
    sys.exit(app.exec())
